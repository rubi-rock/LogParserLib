import csv
import openpyxl
import logging
import os

import os_path_helper
from constants import Headers, IndexedHeaders

def InitializeWorkbook(workbook, create_mode):
    if create_mode:
        row_count = 2
        worksheet = workbook.active
        worksheet.title="Logs compilation"
        worksheet.sheet_properties.tabColor = "1072BA"
        worksheet.cell(row = 1, column=1, value = 'Row count: ')
        worksheet.cell(row = 1, column=2, value = row_count)
        c = 1
        for header in Headers:
            worksheet.cell(row = 2, column = c, value = header)
            c += 1
    else:
        worksheet = workbook.get_sheet_by_name("Logs compilation")
        row_count = worksheet.cell(row = 1, column = 2).value

    return worksheet, row_count


def add_row(worksheet, row_count, row):
    row_count += 1
    for header, value in row.items():
        try:
            worksheet.cell(row = row_count, column = IndexedHeaders[header], value = value)
        except:
            pass
    worksheet.cell(row=1, column=2, value=row_count)
    return row_count


# Never tested...!!!
def WriteParsedLoFileToXSLX(parsed_file, xlsx_file_name):
    try:
        if os.path.isabs(xlsx_file_name):
            xlsx_file_name = os.path.join(os.path.curdir, xlsx_file_name)

        create_mode = not os.path.exists(xlsx_file_name)
        if create_mode:
            workbook = openpyxl.Workbook(encoding='liso-8859-1')
        else:
            workbook = openpyxl.load_workbook(xlsx_file_name)

        worksheet, row_count = InitializeWorkbook(workbook, create_mode)

        # dump the file
        file_name = parsed_file.parsed_file_info.fullname
        row_count = add_row(worksheet, row_count, parsed_file.as_csv_row)
        # dump the sessions
        for session in parsed_file.sessions:
            # dump one session
            row = session.as_csv_row
            row[Headers.file] = file_name
            row_count = add_row(worksheet, row_count, row)
            # dump the session' lines
            for line in session.lines:
                row = line.as_csv_row
                row[Headers.file] = file_name
                row[Headers.context] = row[Headers.context].replace('\t', '    ').replace('\n', '\\n').replace('\r', '\\r')
                row[Headers.session] = session.session_id
                row_count = add_row(worksheet, row_count, row)

        workbook.save(xlsx_file_name)
    except Exception:
        logging.exception('')


def WriteLogFolderParserToXSLX(log_folder_parser, xlsx_file_name=None):
    try:
        if xlsx_file_name is None:
            xlsx_file_name = os_path_helper.generate_file_name(None)
        elif not os.path.isabs(xlsx_file_name):
            xlsx_file_name = os.path.join(os.path.curdir, xlsx_file_name)

        with open(xlsx_file_name, 'w') as csvfile:
            writer = csv.DictWriter(csvfile, delimiter=',', extrasaction='ignore', quoting=csv.QUOTE_ALL,
                                    fieldnames=Headers)
            writer.writeheader()

            # dump the file
            for parsed_file in log_folder_parser.parsed_files:
                file_name = parsed_file.parsed_file_info.fullname
                writer.writerow(parsed_file.as_csv_row)
                # dump the sessions
                for session in parsed_file.sessions:
                    row = session.as_csv_row
                    row[Headers.file] = file_name
                    writer.writerow(row)
                    # dump the lines
                    for line in session.lines:
                        row = line.as_csv_row
                        row[Headers.file] = file_name
                        row[Headers.message] = row[Headers.message].encode('latin-1')       # because of the french text
                        row[Headers.context] = row[Headers.context].encode('latin-1')       # because of the french text
                        row[Headers.session] = session.session_id
                        writer.writerow(row)

        return xlsx_file_name
    except Exception:
        logging.exception('')
